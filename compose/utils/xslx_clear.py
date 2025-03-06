from openpyxl import load_workbook

# 加载 Excel 文件
wb = load_workbook("../template.xlsx")
ws = wb.active

# 删除所有看似空的行（无数据行）
for row in range(ws.max_row, 0, -1):  # 从最后一行开始
    if all(cell.value is None for cell in ws[row]):
        ws.delete_rows(row, 1)

# 保存文件
wb.save("../cleaned_template.xlsx")
print("已清理空行")