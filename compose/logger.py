import yaml
import openpyxl
import time
import matplotlib.pyplot as plt
import os
from datetime import datetime, timedelta
from matplotlib import font_manager
import pickle
class Logger:
    def __init__(self, config, env,plan_path,config_path):
        print("Initializing logger")
        self.path = config["log_path"]
        self.config_path=config_path
        self.plan_path=plan_path
        self.load()
        self.env = env
        self.server_dict = {
            S: {'cores': [], 'GPU': [], 'core_utilization': [], 'GPU_utilization': [], 'ongoing_metatasks': []} for S in
            list(self.env.Server.keys())}
        self.antenna_dict = {
            A: {'beams': [], 'units': []} for A in
            list(self.env.Antennas.keys())}
        self.time_list = []
        self.index = 1
        self.current_task_row=3
        self.success_count=0

        plt.rcParams["font.sans-serif"] = ["SimHei"]

        # font_path = '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc'
        # font_prop = font_manager.FontProperties(fname=font_path)
        #
        # # 配置 Matplotlib 使用该字体
        # plt.rcParams['font.family'] = font_prop.get_name()
        print("Done\n")

    def load(self):
        print(self.path)
        try:
            self.workbook = openpyxl.load_workbook("template.xlsx")
            self.sheet = self.workbook.active
            # 创建表头
        except Exception as e:
            print("Failed to create log file", e)

    def log(self, timestamp, type, res, task,time=None):

        for S in self.server_dict.keys():
            x = self.env.Server[S].idle_cores
            self.server_dict[S]['cores'].append(x)
            y=self.env.Server[S].idle_gpus
            self.server_dict[S]['GPU'].append(y)
            if self.env.Server[S].cpu==0:
                z=0
            else:
                z=x/(self.env.Server[S].cpu)
            self.server_dict[S]['core_utilization'].append(z)
            if self.env.Server[S].gpu==0:
                w=0
            else:
                w=y/(self.env.Server[S].gpu)
            self.server_dict[S]['GPU_utilization'].append(w)
            self.server_dict[S]['ongoing_metatasks'].append(len(self.env.Server[S].ongoing_meta_tasks))
        for A in self.antenna_dict.keys():
            unit_num = self.env.Antennas[A].get_ongoing_unit()
            beam_num = self.env.Antennas[A].ongoing_beam_num
            self.antenna_dict[A]['beams'].append(beam_num)
            self.antenna_dict[A]['units'].append(unit_num)
        self.time_list.append(timestamp)
        resources=task.task_resources


        if type=="req" and res:
            if res=="SUCCESS":
                self.success_count+=1

            self.sheet.cell(row=self.current_task_row, column=1, value=self.index)
            self.sheet.cell(row=self.current_task_row, column=2, value=task.task_name)
            self.sheet.cell(row=self.current_task_row, column=3, value=res)
            self.sheet.cell(row=self.current_task_row, column=4, value=time)
            new_time = datetime.now() + timedelta(seconds=timestamp)
            self.sheet.cell(row=self.current_task_row, column=13, value=new_time)
            temp=0
            for i in range(len(task.task_resources["server"])):
                self.sheet.cell(row=self.current_task_row+i, column=5, value=task.task_resources["server"][i]["metatask"].meta_task_name)
                self.sheet.cell(row=self.current_task_row+i, column=6, value=task.task_resources["server"][i]["id"])
                self.sheet.cell(row=self.current_task_row + i, column=7,
                                value=resources["server"][i]["metatask"].core)
                self.sheet.cell(row=self.current_task_row + i, column=8,
                                value=resources["server"][i]["metatask"].GPU)
                temp=max(temp,i)
            for i in range(len(task.task_resources["antenna"])):
                self.sheet.cell(row=self.current_task_row+i, column=9, value=resources["antenna"][i]["beam"].beam_id)
                self.sheet.cell(row=self.current_task_row+ i, column=10,
                                value=resources["antenna"][i]["antenna_id"])
                self.sheet.cell(row=self.current_task_row + i, column=11,
                                value=len(resources["antenna"][i]["unit_ids"]))
                self.sheet.cell(row=self.current_task_row+i, column=12, value=str(resources["antenna"][i]["unit_ids"]))
                temp = max(temp, i)
            self.current_task_row+=temp+2
            self.index+=1
    def draw(self):
        plt.figure(figsize=(25.6, 14.4))

        plt.subplot(2, 2, 1)
        for key, values in self.server_dict.items():
            plt.plot(self.time_list, values['cores'])
        plt.legend(["SERVER:"+S for S in self.server_dict.keys()])
        plt.xlabel("时间")
        plt.ylabel("剩余核心容量")
        plt.gca().set_ylim(bottom=0)
        plt.title("核心数剩余容量")
        plt.subplot(2, 2, 2)
        for key, values in self.server_dict.items():
            plt.plot(self.time_list, values['GPU'])
        plt.legend([S for S in self.server_dict.keys()])
        plt.xlabel("时间")
        plt.ylabel("剩余GPU")
        plt.gca().set_ylim(bottom=0)
        plt.title("GPU剩余容量")
        plt.subplot(2, 2, 3)
        for key, values in self.server_dict.items():
            plt.plot(self.time_list, values['core_utilization'])
        plt.legend(["SERVER:"+S for S in self.server_dict.keys()])
        plt.xlabel("时间")
        plt.ylabel("可用率")
        plt.ylim(0, 1)
        plt.title("核心可用率")
        plt.subplot(2, 2, 4)
        for key, values in self.server_dict.items():
            plt.plot(self.time_list, values['GPU_utilization'])
        plt.legend(["SERVER:"+S for S in self.server_dict.keys()])
        plt.xlabel("时间")
        plt.ylabel("可用率")
        plt.ylim(0, 1)
        plt.title("GPU可用率")
        plt.rcParams['figure.figsize'] = (12.8, 7.2)

    def draw_ongoing_meta_task(self):
        plt.figure(figsize=(25.6, 14.4))

        for key, values in self.server_dict.items():
            plt.plot(self.time_list, values['ongoing_metatasks'])
        plt.legend([ S for S in self.server_dict.keys()])
        plt.xlabel("时间")
        plt.ylabel("元任务数")
        plt.title("元任务运行数量")

    def draw_antenna(self):
        plt.figure(figsize=(25.6, 14.4))

        plt.subplot(2, 1, 1)
        for key, values in self.antenna_dict.items():
            plt.plot(self.time_list, values['units'])
        plt.legend([A.name for A in self.env.Antennas.values()])
        plt.xlabel("时间")
        plt.ylabel("工作阵元数")
        plt.title("天线阵元使用数")
        plt.subplot(2,1, 2)
        for key, values in self.antenna_dict.items():
            plt.plot(self.time_list, values['beams'])
        plt.legend([A.name for A in self.env.Antennas.values()])
        plt.xlabel("时间")
        plt.ylabel("波束数量")
        plt.title("天线波束数量")

    def conclude(self,runtime,args):
        self.sheet.append([""])
        self.sheet.append(['任务总数',"调度成功数","成功率" ,'环境设置文件', '计划文件',"调度算法","平均时延",'平均任务阵元数'])
        row=str(self.sheet.max_row)
        e1="=AVERAGE(D3:D%s)"%row
        e2="=SUM(K3:K%s)/%d"%(row,self.success_count)
        self.sheet.append([self.index-1,self.success_count,str(f"{self.success_count/(self.index-1):.2%}") ,self.config_path, self.plan_path,self.env.args.method,e1,e2 ])
        self.sheet.append(['当前筛选任务时延'])
        e3="=SUBTOTAL(1,D3:D%s)"%row
        self.sheet.append([e3])

    def save(self):
        dir = time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime())
        os.makedirs(self.path + dir, exist_ok=True)

        self.draw()
        plt.savefig(self.path +dir+ "/utilization" + ".png")

        self.draw_antenna()
        plt.savefig(self.path + dir + "/antenna" + ".png")

        self.draw_ongoing_meta_task()
        plt.savefig(self.path + dir + "/ongoing_metatasks" + ".png")

        self.workbook.save(self.path +dir+ "/log" + ".xlsx")
        picklef=open(self.path +dir+'/pickle_file.pkl','wb')
        pickle.dump(self.env,picklef)